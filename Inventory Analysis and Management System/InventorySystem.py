#installed openpyxl library
import openpyxl

#Load the excel file
excel=openpyxl.load_workbook("C:/Users/lenovo/OneDrive/Desktop/MAIT/3RD SEMESTER/python/learning/project2demofile.xlsx")

#Get the sheet 'page1'
product_list=excel["page1"]

#Dictionary to store total inventory value per supplier
total_value_per_supplier={}

#Dictionary to store products with inventory less than 10
products_under_10_inventory={}

# Dictionary to store number of suppliers per company
products_per_supplier={}

#Loop through all rows
for product_row in range(2,product_list.max_row+1):

    #Get data from cells
    supplier_name=product_list.cell(product_row,4).value
    inventory=product_list.cell(product_row,2).value
    price=product_list.cell(product_row,3).value
    product_num=product_list.cell(product_row,1).value
  
    inventory_value=product_list.cell(product_row,5)

    #Increment supplier product count
    if supplier_name in products_per_supplier: 
        products_per_supplier[supplier_name]+=1
    else:
        print("adding a new supplier ,"+supplier_name)
        products_per_supplier[supplier_name]=1

    #Update supplier inventory value  
    if supplier_name in total_value_per_supplier:
        current=total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name]=current+inventory*price
    else:
        total_value_per_supplier[supplier_name]=price*inventory  

    #Check for low inventory
    if inventory<=10:
        products_under_10_inventory[product_num] = inventory
        print(f"low stock, reorder product number {product_num}")


    #Update inventory value in sheet
    inventory_value.value=inventory * price   
     
#Print the dictionaries
print(products_per_supplier)

print(total_value_per_supplier)

print(products_under_10_inventory)

#Save updated file
excel.save("updated_excel.xlsx")
